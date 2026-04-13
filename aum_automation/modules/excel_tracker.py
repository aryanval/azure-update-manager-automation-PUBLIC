"""
Excel Tracker - Manages patch tracking Excel files
Maintains historical records even if VMs are deleted mid-cycle
"""

import os
import logging
from datetime import datetime
from typing import Dict, List, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelTracker:
    """
    Manages Excel files for patch tracking.
    
    Key features:
    - Reads template structure from existing file
    - Creates new sheets for each patching cycle
    - Maintains historical records (VMs persist even if deleted)
    - Updates patch status for all VMs
    """
    
    # Column headers based on template structure
    COLUMN_HEADERS = [
        'VM Name',
        'Resource Group',
        'OS Type',
        'Initial State',
        'Patches Available',
        'Patches Applied',
        'Patches Excluded',
        'Status',
        'Error Details',
        'Action Taken',
        'Last Update',
        'Notes'
    ]
    
    # Status colors
    STATUS_COLORS = {
        'Success': '90EE90',      # Light green
        'Failed': 'FFB6C1',       # Light red
        'In Progress': 'FFD700',  # Gold
        'Skipped': 'D3D3D3',      # Light gray
        'Manual Review': 'FFA500', # Orange
        'Not Started': 'FFFFFF'   # White
    }
    
    def __init__(self, config):
        self.config = config
        self.excel_config = config.get_excel_config()
        self.template_path = self.excel_config.get('template_path')
        self.output_dir = self.excel_config.get('output_directory', './reports')
        self.preserve_deleted = self.excel_config.get('preserve_deleted_vms', True)
        
        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Cache for current workbook
        self.current_workbook: Optional[Workbook] = None
        self.current_file_path: Optional[str] = None
    
    def get_cycle_name(self) -> str:
        """
        Get current patching cycle name (based on month/year).
        Microsoft Patch Tuesday: 2nd Tuesday of each month
        """
        now = datetime.now()
        return f"{now.strftime('%b %Y')}"  # e.g., "Apr 2026"
    
    def get_output_file_path(self, tenant_name: str) -> str:
        """Get output file path for current cycle"""
        cycle_name = self.get_cycle_name()
        filename = f"Patching_{tenant_name}_{cycle_name.replace(' ', '_')}.xlsx"
        return os.path.join(self.output_dir, filename)
    
    def load_or_create_workbook(self, tenant_name: str) -> Workbook:
        """
        Load existing workbook for this cycle or create new one from template.
        
        Args:
            tenant_name: Name of the tenant (as defined in config.yaml)
        
        Returns:
            Workbook instance
        """
        output_path = self.get_output_file_path(tenant_name)
        
        # If file exists for this cycle, load it
        if os.path.exists(output_path):
            logger.info(f"Loading existing cycle workbook: {output_path}")
            wb = load_workbook(output_path)
            self.current_workbook = wb
            self.current_file_path = output_path
            return wb
        
        # Create new workbook from template structure
        logger.info(f"Creating new workbook from template for {tenant_name}")
        wb = Workbook()
        
        # Create sheets based on tenant
        sheet_names = self._get_sheet_names_for_tenant(tenant_name)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets with headers
        for sheet_name in sheet_names:
            ws = wb.create_sheet(sheet_name)
            self._setup_sheet_headers(ws)
        
        # Save initial workbook
        wb.save(output_path)
        logger.info(f"Created new workbook: {output_path}")
        
        self.current_workbook = wb
        self.current_file_path = output_path
        return wb
    
    def _get_sheet_names_for_tenant(self, tenant_name: str) -> List[str]:
        """
        Get sheet names for a tenant.
        Dynamically generates sheets based on tenant config.
        If the tenant config specifies avd_enabled=true, an AVD sheet is added.
        """
        tenant_cfg = {}
        try:
            tenant_cfg = self.config.get_tenant_config(tenant_name)
        except Exception:
            pass

        sheets = [f"{tenant_name} Servers"]
        if tenant_cfg.get("avd_enabled", False):
            sheets.append(f"{tenant_name} AVDs")
        return sheets
    
    def _setup_sheet_headers(self, worksheet):
        """Setup headers for a worksheet"""
        # Write headers
        for idx, header in enumerate(self.COLUMN_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        column_widths = {
            'VM Name': 30,
            'Resource Group': 25,
            'OS Type': 15,
            'Initial State': 15,
            'Patches Available': 15,
            'Patches Applied': 15,
            'Patches Excluded': 15,
            'Status': 15,
            'Error Details': 40,
            'Action Taken': 20,
            'Last Update': 20,
            'Notes': 30
        }
        
        for idx, header in enumerate(self.COLUMN_HEADERS, start=1):
            worksheet.column_dimensions[get_column_letter(idx)].width = column_widths.get(header, 15)
        
        # Freeze top row
        worksheet.freeze_panes = 'A2'
    
    def find_vm_row(self, worksheet, vm_name: str) -> Optional[int]:
        """
        Find the row number for a VM in the worksheet.
        
        Returns:
            Row number if found, None otherwise
        """
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=row, column=1).value == vm_name:
                return row
        return None
    
    def update_vm_status(
        self, 
        tenant_name: str,
        sheet_type: str,  # 'Servers' or 'AVDs'
        vm_data: Dict
    ):
        """
        Update or add VM status in the appropriate sheet.
        
        Args:
            tenant_name: Tenant name (Essen, ECares, IHV)
            sheet_type: 'Servers' or 'AVDs'
            vm_data: Dictionary with VM information
        """
        wb = self.load_or_create_workbook(tenant_name)
        sheet_name = f"{tenant_name} {sheet_type}"
        
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found, creating it")
            ws = wb.create_sheet(sheet_name)
            self._setup_sheet_headers(ws)
        else:
            ws = wb[sheet_name]
        
        vm_name = vm_data.get('vm_name', '')
        
        # Find existing row or create new one
        row_num = self.find_vm_row(ws, vm_name)
        
        if row_num is None:
            # Add new row
            row_num = ws.max_row + 1
            logger.info(f"Adding new VM: {vm_name} to row {row_num}")
        else:
            logger.info(f"Updating existing VM: {vm_name} at row {row_num}")
        
        # Update cells
        ws.cell(row=row_num, column=1).value = vm_name
        ws.cell(row=row_num, column=2).value = vm_data.get('resource_group', '')
        ws.cell(row=row_num, column=3).value = vm_data.get('os_type', '')
        ws.cell(row=row_num, column=4).value = vm_data.get('initial_state', '')
        ws.cell(row=row_num, column=5).value = vm_data.get('patches_available', 0)
        ws.cell(row=row_num, column=6).value = vm_data.get('patches_applied', 0)
        ws.cell(row=row_num, column=7).value = vm_data.get('patches_excluded', 0)
        
        status = vm_data.get('status', 'Not Started')
        ws.cell(row=row_num, column=8).value = status
        
        # Apply status color
        if status in self.STATUS_COLORS:
            status_cell = ws.cell(row=row_num, column=8)
            status_cell.fill = PatternFill(
                start_color=self.STATUS_COLORS[status],
                end_color=self.STATUS_COLORS[status],
                fill_type='solid'
            )
        
        ws.cell(row=row_num, column=9).value = vm_data.get('error_details', '')
        ws.cell(row=row_num, column=10).value = vm_data.get('action_taken', '')
        ws.cell(row=row_num, column=11).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=row_num, column=12).value = vm_data.get('notes', '')
        
        # Save workbook
        wb.save(self.current_file_path)
        logger.debug(f"Saved workbook: {self.current_file_path}")
    
    def get_existing_vms(self, tenant_name: str, sheet_type: str) -> List[str]:
        """
        Get list of VMs already in the sheet (for historical tracking).
        
        Args:
            tenant_name: Tenant name
            sheet_type: 'Servers' or 'AVDs'
        
        Returns:
            List of VM names
        """
        wb = self.load_or_create_workbook(tenant_name)
        sheet_name = f"{tenant_name} {sheet_type}"
        
        if sheet_name not in wb.sheetnames:
            return []
        
        ws = wb[sheet_name]
        vm_names = []
        
        for row in range(2, ws.max_row + 1):
            vm_name = ws.cell(row=row, column=1).value
            if vm_name:
                vm_names.append(vm_name)
        
        return vm_names
    
    def mark_vm_as_deleted(self, tenant_name: str, sheet_type: str, vm_name: str):
        """
        Mark a VM as deleted/not found (for historical tracking).
        Preserves the record but updates status.
        
        Args:
            tenant_name: Tenant name
            sheet_type: 'Servers' or 'AVDs'
            vm_name: VM name
        """
        if not self.preserve_deleted:
            return
        
        vm_data = {
            'vm_name': vm_name,
            'status': 'Not Found',
            'action_taken': 'VM deleted or moved',
            'notes': f'VM not found in Azure inventory as of {datetime.now().strftime("%Y-%m-%d")}'
        }
        
        self.update_vm_status(tenant_name, sheet_type, vm_data)
        logger.info(f"Marked VM as deleted: {vm_name}")
    
    def create_summary_report(self, tenant_name: str) -> Dict:
        """
        Create a summary report from the current cycle workbook.
        
        Returns:
            Dictionary with summary statistics
        """
        wb = self.load_or_create_workbook(tenant_name)
        
        summary = {
            'tenant': tenant_name,
            'cycle': self.get_cycle_name(),
            'total_vms': 0,
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'manual_review': 0,
            'in_progress': 0,
            'not_started': 0,
            'total_patches_applied': 0,
            'total_patches_excluded': 0
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in range(2, ws.max_row + 1):
                vm_name = ws.cell(row=row, column=1).value
                if not vm_name:
                    continue
                
                summary['total_vms'] += 1
                
                status = ws.cell(row=row, column=8).value or 'Not Started'
                
                if 'Success' in status:
                    summary['success'] += 1
                elif 'Failed' in status:
                    summary['failed'] += 1
                elif 'Skipped' in status:
                    summary['skipped'] += 1
                elif 'Manual' in status:
                    summary['manual_review'] += 1
                elif 'Progress' in status:
                    summary['in_progress'] += 1
                else:
                    summary['not_started'] += 1
                
                # Count patches
                patches_applied = ws.cell(row=row, column=6).value or 0
                patches_excluded = ws.cell(row=row, column=7).value or 0
                
                summary['total_patches_applied'] += patches_applied if isinstance(patches_applied, int) else 0
                summary['total_patches_excluded'] += patches_excluded if isinstance(patches_excluded, int) else 0
        
        return summary
    
    def close(self):
        """Close the current workbook"""
        if self.current_workbook:
            self.current_workbook.close()
            self.current_workbook = None
            self.current_file_path = None
